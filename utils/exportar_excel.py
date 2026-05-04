from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font


def exportar_treeview(tree, solo_seleccion=False):

    if solo_seleccion:
        filas = tree.selection()
    else:
        filas = tree.get_children()

    if not filas:
        messagebox.showerror("Aviso", "No hay datos para exportar")
        return

    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
    filetypes=[("Excel File", ".xlsx")],
    )

    if not ruta:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'Datos'


    # Encabezados
    columnas = tree['columns']

    for col_num, col in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=col).font = Font(bold=True)
        cell.font = Font(bold=True)

    # Datos
    for row_num, fila in enumerate(filas, 2):
        valores = tree.item(fila)['values']

        for col_num, valor in enumerate(valores, 1):
            ws.cell(row=row_num, column=col_num, value=valor)

    # Auto ajuste columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(ruta)

    messagebox.showinfo("Exito", "Exportar datos correctamente")